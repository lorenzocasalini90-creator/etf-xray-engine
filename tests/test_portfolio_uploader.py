"""Tests for portfolio file uploader."""

import io

import pytest

from src.dashboard.components.portfolio_uploader import (
    generate_template_xlsx,
    parse_portfolio_file,
    _parse_amount,
)


class TestParseAmount:
    def test_simple_integer(self):
        assert _parse_amount("30000") == 30000.0

    def test_with_euro_sign(self):
        assert _parse_amount("€30000") == 30000.0

    def test_single_dot_ambiguous_treated_as_decimal(self):
        """Single dot with 3 digits after: ambiguous — treat as decimal for safety."""
        assert _parse_amount("30.000") == 30.0
        assert _parse_amount("13.313") == 13.313
        assert _parse_amount("9.268") == 9.268

    def test_european_decimal(self):
        assert _parse_amount("30.000,50") == 30000.50

    def test_us_thousands(self):
        assert _parse_amount("30,000") == 30000.0

    def test_us_decimal(self):
        assert _parse_amount("30,000.50") == 30000.50

    def test_euro_with_spaces(self):
        assert _parse_amount("€ 30 000") == 30000.0

    def test_empty_returns_none(self):
        assert _parse_amount("") is None

    def test_invalid_returns_none(self):
        assert _parse_amount("abc") is None

    def test_european_full_format(self):
        """13.313,125 → 13313.125 (European: dot=thousands, comma=decimal)"""
        assert _parse_amount("13.313,125") == 13313.125

    def test_comma_thousands_three_digits(self):
        """13313,125 — only comma, exactly 3 digits after → thousands."""
        assert _parse_amount("13313,125") == 13313125.0

    def test_dollar_sign(self):
        assert _parse_amount("$30000") == 30000.0

    def test_dollar_with_comma(self):
        assert _parse_amount("$30,000.50") == 30000.50

    def test_numeric_int_passthrough(self):
        assert _parse_amount(30000) == 30000.0

    def test_numeric_float_passthrough(self):
        assert _parse_amount(13313.125) == 13313.125

    def test_comma_decimal_two_digits(self):
        """13313,12 → 13313.12 (decimal comma, 2 digits)"""
        assert _parse_amount("13313,12") == 13313.12

    def test_dot_decimal_two_digits(self):
        """13313.12 → 13313.12 (decimal dot, 2 digits)"""
        assert _parse_amount("13313.12") == 13313.12

    def test_string_float_decimal_not_thousands(self):
        """'13313.125' as string — decimal, NOT thousands × 1000."""
        assert _parse_amount("13313.125") == 13313.125

    def test_multi_dot_eu_thousands(self):
        """'1.313.000' — multiple dot-groups of 3 → unambiguous EU thousands."""
        assert _parse_amount("1.313.000") == 1313000.0


class TestGenerateTemplate:
    def test_generates_bytes(self):
        data = generate_template_xlsx()
        assert isinstance(data, bytes)
        assert len(data) > 100

    def test_readable_by_openpyxl(self):
        import openpyxl
        data = generate_template_xlsx()
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.cell(1, 1).value == "Ticker/ISIN"
        assert ws.cell(2, 1).value == "CSPX"
        assert ws.cell(2, 2).value == 30000


class TestParseCSV:
    def _make_csv(self, content: str) -> io.StringIO:
        return io.StringIO(content)

    def test_basic_csv(self):
        csv = self._make_csv("Ticker/ISIN,Importo EUR\nCSPX,30000\nSWDA,40000\n")
        positions, errors = parse_portfolio_file(csv, filename="test.csv")
        assert len(positions) == 2
        assert positions[0] == {"ticker": "CSPX", "capital": 30000.0}

    def test_semicolon_csv(self):
        csv = self._make_csv("Ticker/ISIN;Importo EUR\nCSPX;30000\nSWDA;40000\n")
        positions, errors = parse_portfolio_file(csv, filename="test.csv")
        assert len(positions) == 2

    def test_euro_amounts(self):
        csv = self._make_csv("ticker;importo\nCSPX;€30.000,00\nSWDA;€40.000,50\n")
        positions, errors = parse_portfolio_file(csv, filename="test.csv")
        assert positions[0]["capital"] == 30000.0

    def test_duplicate_sums(self):
        csv = self._make_csv("ticker,importo\nCSPX,10000\nCSPX,20000\n")
        positions, errors = parse_portfolio_file(csv, filename="test.csv")
        assert len(positions) == 1
        assert positions[0]["capital"] == 30000.0
        assert any("sommati" in e for e in errors)

    def test_missing_columns_error(self):
        csv = self._make_csv("col1,col2\na,b\n")
        positions, errors = parse_portfolio_file(csv, filename="test.csv")
        assert len(positions) == 0
        assert any("Colonne non trovate" in e for e in errors)

    def test_invalid_amount_row(self):
        csv = self._make_csv("ticker,importo\nCSPX,30000\nSWDA,abc\n")
        positions, errors = parse_portfolio_file(csv, filename="test.csv")
        assert len(positions) == 1
        assert any("non valido" in e for e in errors)


class TestParseExcel:
    def test_template_roundtrip(self):
        data = generate_template_xlsx()
        positions, errors = parse_portfolio_file(io.BytesIO(data), filename="test.xlsx")
        assert len(positions) == 3
        assert positions[0] == {"ticker": "CSPX", "capital": 30000.0}
        assert positions[1] == {"ticker": "SWDA", "capital": 40000.0}
        assert positions[2] == {"ticker": "VWCE", "capital": 15000.0}
